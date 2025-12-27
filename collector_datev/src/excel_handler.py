from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .plz_filter import PlzFilter, matches_filter


@dataclass
class PlzEntry:
    """Represents a single PLZ row from the Excel file."""

    row_number: int
    plz: str
    city: str


@dataclass
class LocationGroup:
    """A group of PLZ entries for the same city."""

    city: str
    entries: list[PlzEntry]


def load_pending_locations(path: Path, plz_filter: PlzFilter | None = None) -> list[LocationGroup]:
    """Load all unprocessed PLZ entries grouped by city.

    Only returns entries where column D (Verarbeitung_Datum) is empty.

    Args:
        path: Path to the Excel file
        plz_filter: If specified, only load PLZ matching the filter
    """
    wb = load_workbook(path, read_only=True)
    ws = wb.active

    city_groups: dict[str, list[PlzEntry]] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=6), start=2):
        plz = row[0].value
        city = row[1].value
        processed_date = row[3].value if len(row) > 3 else None

        if not plz or not city:
            continue

        if processed_date:
            continue

        plz_str = str(plz).zfill(5)

        # Filter by PLZ if specified
        if plz_filter is not None and not matches_filter(plz_str, plz_filter):
            continue

        if city not in city_groups:
            city_groups[city] = []

        city_groups[city].append(PlzEntry(row_number=row_idx, plz=plz_str, city=city))

    wb.close()

    return [LocationGroup(city=city, entries=entries) for city, entries in city_groups.items()]


def update_plz_status(
    path: Path,
    row_number: int,
    count: int,
    error: str | None = None,
) -> None:
    """Update the status columns for a single PLZ row.

    Args:
        path: Path to the Excel file
        row_number: The row number to update (1-indexed)
        count: Number of Steuerberater found
        error: Error message if any (written to column F)
    """
    wb = load_workbook(path)
    ws: Worksheet = wb.active

    today = datetime.now().strftime("%Y-%m-%d")
    ws.cell(row_number, 4, today)
    ws.cell(row_number, 5, count)

    if error:
        ws.cell(row_number, 6, error)

    wb.save(path)
    wb.close()


def get_progress_stats(path: Path) -> dict[str, int]:
    """Get statistics about processing progress."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active

    total = 0
    processed = 0
    errors = 0

    for row in ws.iter_rows(min_row=2, max_col=6):
        plz = row[0].value
        if not plz:
            continue

        total += 1

        if len(row) > 3 and row[3].value:
            processed += 1

        if len(row) > 5 and row[5].value:
            errors += 1

    wb.close()

    return {
        "total": total,
        "processed": processed,
        "pending": total - processed,
        "errors": errors,
    }
