"""Flask CLI Commands.

Custom commands for database initialization and blacklist management.
"""

from __future__ import annotations

import click
from flask import Flask
from flask.cli import with_appcontext

from app import db
from app.models import Category, Domain, Plz, Rechtsform, PlzCollector, Kammer, Kanzlei, Steuerberater


def register_commands(app: Flask):
    """Register all CLI commands with the Flask app."""
    app.cli.add_command(init_db_command)
    app.cli.add_command(import_blacklist_command)
    app.cli.add_command(export_blacklist_command)
    app.cli.add_command(seed_command)
    app.cli.add_command(seed_categories_command)
    app.cli.add_command(import_plz_command)
    app.cli.add_command(stats_plz_command)
    # BStBK Collector commands
    app.cli.add_command(seed_rechtsformen_command)
    app.cli.add_command(init_plz_collector_command)
    app.cli.add_command(stats_bstbk_command)


@click.command("init-db")
@with_appcontext
def init_db_command():
    """Initialize the database (create all tables)."""
    db.create_all()
    click.echo("Datenbank initialisiert.")


@click.command("seed-categories")
@with_appcontext
def seed_categories_command():
    """Create default categories if they don't exist."""
    count = Category.seed_defaults()
    if count > 0:
        click.echo(f"{count} Standard-Kategorien erstellt.")
    else:
        click.echo("Alle Standard-Kategorien existieren bereits.")


@click.command("import-blacklist")
@click.argument("filepath", default="data/domain_blacklist.txt")
@with_appcontext
def import_blacklist_command(filepath: str):
    """Import blacklist from TXT file.

    FILEPATH: Path to the blacklist file (default: data/domain_blacklist.txt)
    """
    from pathlib import Path
    from flask import current_app

    # Ensure categories exist first
    Category.seed_defaults()

    # Resolve path relative to project root
    if not filepath.startswith("/"):
        filepath = current_app.config["PROJECT_ROOT"] / filepath
    else:
        filepath = Path(filepath)

    if not filepath.exists():
        click.echo(f"Datei nicht gefunden: {filepath}")
        return

    count = 0
    current_category_slug = "unsortiert"

    # Category mapping for common names in TXT file
    category_mappings = {
        "email-provider": "email-provider",
        "e-mail provider": "email-provider",
        "hosting": "hosting",
        "hosting provider": "hosting",
        "verzeichnis": "verzeichnis",
        "verzeichnisse": "verzeichnis",
        "steuerberater-verzeichnisse": "verzeichnis",
        "social-media": "social-media",
        "unsortiert": "unsortiert",
    }

    with open(filepath, "r") as f:
        for line in f:
            line = line.strip()

            # Skip empty lines
            if not line:
                continue

            # Category header
            if line.startswith("#"):
                category_name = line[1:].strip().lower()
                current_category_slug = category_mappings.get(category_name, "unsortiert")
                continue

            # Domain line
            domain = line.lower()
            existing = Domain.query.filter_by(domain=domain).first()
            if not existing:
                # Find category by slug
                category = Category.query.filter_by(slug=current_category_slug).first()

                new_domain = Domain(
                    domain=domain,
                    category_id=category.id if category else None,
                    created_by="import",
                )
                db.session.add(new_domain)
                count += 1

    db.session.commit()
    click.echo(f"{count} Domains importiert.")


@click.command("export-blacklist")
@click.argument("filepath", default="data/domain_blacklist.txt")
@with_appcontext
def export_blacklist_command(filepath: str):
    """Export blacklist to TXT file.

    FILEPATH: Path to the output file (default: data/domain_blacklist.txt)
    """
    from pathlib import Path
    from flask import current_app

    # Resolve path relative to project root
    if not filepath.startswith("/"):
        filepath = current_app.config["PROJECT_ROOT"] / filepath
    else:
        filepath = Path(filepath)

    # Get all domains ordered by category and domain name
    domains = (
        Domain.query
        .outerjoin(Category)
        .order_by(Category.sort_order, Domain.domain)
        .all()
    )

    with open(filepath, "w") as f:
        current_category_id = -1  # Sentinel to detect first category
        for domain in domains:
            cat_id = domain.category_id or 0
            if cat_id != current_category_id:
                if current_category_id != -1:
                    f.write("\n")
                category_name = domain.category_slug if domain.category_rel else "unsortiert"
                f.write(f"# {category_name}\n")
                current_category_id = cat_id
            f.write(f"{domain.domain}\n")

    click.echo(f"{len(domains)} Domains exportiert nach {filepath}.")


@click.command("seed")
@with_appcontext
def seed_command():
    """Seed the database with initial data (categories + existing blacklist)."""
    from pathlib import Path
    from flask import current_app

    # First, create default categories
    click.echo("Erstelle Standard-Kategorien...")
    count = Category.seed_defaults()
    click.echo(f"  {count} Kategorien erstellt.")

    # Import blacklist if it exists
    blacklist_path = current_app.config["PROJECT_ROOT"] / "data" / "domain_blacklist.txt"
    if blacklist_path.exists():
        click.echo("Importiere bestehende Blacklist...")
        # Use the import command logic inline
        import_blacklist_command.main(["data/domain_blacklist.txt"], standalone_mode=False)
    else:
        click.echo("Keine bestehende Blacklist gefunden.")

    click.echo("Seed abgeschlossen.")


@click.command("import-plz")
@click.argument("excel_path", default="data/plz_de.xlsx")
@with_appcontext
def import_plz_command(excel_path: str):
    """Import PLZ data from Excel file into database.

    EXCEL_PATH: Path to the Excel file (default: data/plz_de.xlsx)
    """
    from datetime import datetime
    from pathlib import Path

    from flask import current_app
    from openpyxl import load_workbook

    def parse_date(value):
        """Parse date from Excel - handles datetime, date, and string formats."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        if hasattr(value, "year"):  # date object
            return datetime(value.year, value.month, value.day)
        # Try to parse string
        if isinstance(value, str):
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S"):
                try:
                    return datetime.strptime(value, fmt)
                except ValueError:
                    continue
        return None

    # Resolve path relative to project root
    if not excel_path.startswith("/"):
        excel_path = current_app.config["PROJECT_ROOT"] / excel_path
    else:
        excel_path = Path(excel_path)

    if not excel_path.exists():
        click.echo(f"Datei nicht gefunden: {excel_path}")
        return

    click.echo(f"Importiere PLZ aus {excel_path}...")

    try:
        wb = load_workbook(excel_path, read_only=True)
    except Exception as e:
        click.echo(f"Fehler beim Öffnen der Excel-Datei: {e}")
        return

    ws = wb.active

    imported = 0
    updated = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, max_col=6):
        plz_val = row[0].value
        city = row[1].value
        # Column C is unused
        processed_date = parse_date(row[3].value if len(row) > 3 else None)
        count = row[4].value if len(row) > 4 else None
        error = row[5].value if len(row) > 5 else None

        if not plz_val or not city:
            skipped += 1
            continue

        plz_str = str(plz_val).zfill(5)

        # Check if PLZ already exists
        existing = Plz.query.filter_by(plz=plz_str).first()

        if existing:
            # Update existing entry
            existing.city = city
            existing.processed_at = processed_date
            existing.result_count = count
            existing.error_message = error
            updated += 1
        else:
            # Create new entry
            plz_entry = Plz(
                plz=plz_str,
                city=city,
                processed_at=processed_date,
                result_count=count,
                error_message=error,
            )
            db.session.add(plz_entry)
            imported += 1

    db.session.commit()
    wb.close()

    click.echo(f"Import abgeschlossen:")
    click.echo(f"  Neu importiert: {imported}")
    click.echo(f"  Aktualisiert:   {updated}")
    click.echo(f"  Übersprungen:   {skipped}")
    click.echo(f"  Gesamt in DB:   {Plz.query.count()}")


@click.command("stats-plz")
@with_appcontext
def stats_plz_command():
    """Show PLZ processing statistics."""
    total = Plz.query.count()

    if total == 0:
        click.echo("Keine PLZ in der Datenbank. Führe 'flask import-plz' aus.")
        return

    processed = Plz.query.filter(Plz.processed_at.isnot(None)).count()
    pending = total - processed
    errors = Plz.query.filter(Plz.error_message.isnot(None)).count()
    total_found = db.session.query(db.func.sum(Plz.result_count)).scalar() or 0

    click.echo("PLZ-Statistiken:")
    click.echo(f"  Gesamt PLZ:          {total:,}")
    click.echo(f"  Verarbeitet:         {processed:,} ({processed*100//total}%)")
    click.echo(f"  Ausstehend:          {pending:,} ({pending*100//total}%)")
    click.echo(f"  Mit Fehlern:         {errors:,}")
    click.echo(f"  Steuerberater total: {total_found:,}")


# ============================================================================
# BStBK Scraper Commands
# ============================================================================


@click.command("seed-rechtsformen")
@with_appcontext
def seed_rechtsformen_command():
    """Seed the Rechtsform table with default values."""
    count = Rechtsform.seed_defaults()
    if count > 0:
        click.echo(f"{count} Rechtsformen erstellt.")
    else:
        click.echo("Alle Rechtsformen existieren bereits.")


@click.command("init-plz-collector")
@click.argument("collector_type", default="bstbk")
@click.option("--force", is_flag=True, help="Vorinitialisierung erzwingen (normalerweise nicht nötig)")
@with_appcontext
def init_plz_collector_command(collector_type: str, force: bool):
    """Initialize PLZ tracking for a collector type from the plz table.

    HINWEIS: Die Vorinitialisierung ist OPTIONAL!
    PLZ-Einträge werden automatisch on-demand erstellt, wenn der Collector läuft.

    Verwende diesen Befehl nur, wenn du alle PLZ vorab initialisieren möchtest
    (z.B. für Statistik-Anzeige vor dem ersten Lauf).

    COLLECTOR_TYPE: The collector type to initialize (default: bstbk)
    """
    if not force:
        click.echo("⚠️  HINWEIS: Die Vorinitialisierung ist normalerweise nicht nötig!")
        click.echo("   PLZ-Einträge werden automatisch on-demand erstellt.")
        click.echo("")
        click.echo("   Verwende --force, um trotzdem alle PLZ zu initialisieren.")
        click.echo("   Beispiel: flask init-plz-collector bstbk --force")
        return

    # Check if plz table has entries
    plz_count = Plz.query.count()
    if plz_count == 0:
        click.echo("Keine PLZ in der Datenbank. Führe zuerst 'flask import-plz' aus.")
        return

    # Initialize plz_collector entries
    count = PlzCollector.init_from_plz_table(collector_type)
    db.session.commit()

    if count > 0:
        click.echo(f"{count} PLZ-Einträge für Collector '{collector_type}' initialisiert.")
    else:
        click.echo(f"Alle PLZ für Collector '{collector_type}' existieren bereits.")

    # Show stats
    stats = PlzCollector.get_stats(collector_type)
    click.echo(f"  Gesamt:      {stats['total']:,}")
    click.echo(f"  Verarbeitet: {stats['processed']:,}")
    click.echo(f"  Ausstehend:  {stats['pending']:,}")


@click.command("stats-bstbk")
@with_appcontext
def stats_bstbk_command():
    """Show BStBK collector statistics."""
    click.echo("BStBK Collector Statistiken:")
    click.echo("─" * 40)

    # Total PLZ from reference table
    total_plz = Plz.query.count()

    if total_plz == 0:
        click.echo("Keine PLZ in der Datenbank. Führe 'flask import-plz' aus.")
        return

    # PLZ Stats from plz_collector (on-demand entries)
    collector_stats = PlzCollector.get_stats("bstbk")
    processed = collector_stats["processed"]
    errors = collector_stats["errors"]
    pending = total_plz - processed

    pct = processed * 100 // total_plz if total_plz > 0 else 0
    click.echo(f"PLZ-Fortschritt:")
    click.echo(f"  Gesamt:      {total_plz:,} (aus PLZ-Referenztabelle)")
    click.echo(f"  Verarbeitet: {processed:,} ({pct}%)")
    click.echo(f"  Ausstehend:  {pending:,}")
    click.echo(f"  Mit Fehlern: {errors:,}")

    click.echo("")

    # Kanzlei & Steuerberater Stats
    kanzlei_count = Kanzlei.query.count()
    stb_count = Steuerberater.query.count()
    kammer_count = Kammer.query.count()

    click.echo(f"Datenbestand:")
    click.echo(f"  Kammern:        {kammer_count:,}")
    click.echo(f"  Kanzleien:      {kanzlei_count:,}")
    click.echo(f"  Steuerberater:  {stb_count:,}")

    # Show top Kammern if any exist
    if kammer_count > 0:
        click.echo("")
        click.echo("Top 5 Kammern nach Kanzleien:")
        top_kammern = (
            db.session.query(
                Kammer.name,
                db.func.count(Kanzlei.id).label("count")
            )
            .outerjoin(Kanzlei)
            .group_by(Kammer.id)
            .order_by(db.desc("count"))
            .limit(5)
            .all()
        )
        for name, count in top_kammern:
            click.echo(f"  {name}: {count:,}")
